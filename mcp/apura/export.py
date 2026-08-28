"""Exportação XLSX e HTML a partir dos dados das consultas (Fato / Acervo / Clima)."""
from __future__ import annotations

import html as html_module
import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _flatten_rows(res: dict[str, Any], tool: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    linhas = res.get("linhas")
    if isinstance(linhas, list):
        for row in linhas:
            if isinstance(row, dict):
                r = dict(row)
                r["_consulta"] = tool
                rows.append(r)
    itens = res.get("itens")
    if isinstance(itens, list):
        for row in itens:
            if isinstance(row, dict):
                r = dict(row)
                r["_consulta"] = tool
                rows.append(r)
    series = res.get("series")
    if isinstance(series, list):
        for block in series:
            if not isinstance(block, dict):
                continue
            ano = block.get("ano")
            for row in block.get("linhas") or []:
                if isinstance(row, dict):
                    r = dict(row)
                    r["_consulta"] = tool
                    r["_ano_serie"] = ano
                    rows.append(r)
    for key in ("acervo_a", "acervo_b"):
        sub = res.get(key)
        if isinstance(sub, dict):
            for row in sub.get("itens") or []:
                if isinstance(row, dict):
                    r = dict(row)
                    r["_consulta"] = f"{tool}:{key}"
                    rows.append(r)
    return rows


def _classificar_tool(tool: str) -> str:
    t = (tool or "").lower()
    if "acervo" in t or "ficha" in t:
        return "acervo"
    if "clima" in t:
        return "clima"
    return "fato"


def _linhas_por_camada(dados: dict[str, Any] | None) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {"fato": [], "acervo": [], "clima": []}
    if not dados:
        return out
    for tr in dados.get("tool_results", []):
        tool = tr.get("tool", "")
        res = tr.get("result")
        if not isinstance(res, dict):
            continue
        camada = _classificar_tool(tool)
        out[camada].extend(_flatten_rows(res, tool))
    return out


def _sheet_from_rows(ws, rows: list[dict[str, Any]], titulo_vazio: str) -> None:
    if not rows:
        ws.append([titulo_vazio])
        return
    keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                keys.append(k)
    ws.append(keys)
    for row in rows:
        ws.append([row.get(k, "") for k in keys])
    for i, _ in enumerate(keys, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def exportar_xlsx(dados: dict[str, Any] | None, titulo: str = "Apura") -> bytes:
    camadas = _linhas_por_camada(dados)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Fato"
    _sheet_from_rows(ws0, camadas["fato"], "Sem dados tabulares de urna/contexto nesta resposta")

    ws1 = wb.create_sheet("Acervo")
    _sheet_from_rows(ws1, camadas["acervo"], "Sem trechos de acervo nesta resposta")

    ws2 = wb.create_sheet("Clima")
    _sheet_from_rows(ws2, camadas["clima"], "Sem itens de clima nesta resposta")

    meta = wb.create_sheet("Meta")
    meta.append(["titulo", titulo])
    meta.append(["gerado_em", datetime.now(timezone.utc).isoformat()])
    meta.append(["fato_linhas", len(camadas["fato"])])
    meta.append(["acervo_linhas", len(camadas["acervo"])])
    meta.append(["clima_linhas", len(camadas["clima"])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _esc(text: str) -> str:
    return html_module.escape(text, quote=True)


def _inline_md(text: str) -> str:
    s = _esc(text)
    s = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", s)
    s = re.sub(
        r"\[([^\]]+)\]\((https?://[^)\s]+)\)",
        r'<a href="\2" target="_blank" rel="noopener noreferrer">\1</a>',
        s,
    )
    s = re.sub(
        r"(https?://[^\s<]{80,})",
        r'<a href="\1" target="_blank" rel="noopener noreferrer">abrir link</a>',
        s,
    )
    return s


def _md_block(text: str) -> str:
    if not text.strip():
        return ""
    out: list[str] = []
    in_ul = False

    def close_ul() -> None:
        nonlocal in_ul
        if in_ul:
            out.append("</ul>")
            in_ul = False

    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            close_ul()
            continue
        if stripped.startswith("### "):
            close_ul()
            out.append(f"<h4>{_inline_md(stripped[4:])}</h4>")
        elif stripped.startswith("## "):
            close_ul()
            out.append(f"<h3>{_inline_md(stripped[3:])}</h3>")
        elif stripped.startswith("- ") or stripped.startswith("* "):
            if not in_ul:
                out.append("<ul>")
                in_ul = True
            out.append(f"<li>{_inline_md(stripped[2:])}</li>")
        else:
            close_ul()
            out.append(f"<p>{_inline_md(stripped)}</p>")
    close_ul()
    return "\n".join(out)


def _secoes_from_md(text: str) -> list[tuple[str, str]]:
    parts = re.split(r"^### (.+)$", text, flags=re.MULTILINE)
    if len(parts) <= 1:
        return [("Análise", text.strip())]
    secoes: list[tuple[str, str]] = []
    intro = parts[0].strip()
    if intro:
        secoes.append(("Introdução", intro))
    for i in range(1, len(parts), 2):
        titulo = parts[i].strip()
        corpo = parts[i + 1].strip() if i + 1 < len(parts) else ""
        secoes.append((titulo, corpo))
    return secoes


def _classe_secao(titulo: str) -> str:
    t = titulo.lower()
    if "fato" in t:
        return "sec-fato"
    if "programa" in t or "acervo" in t:
        return "sec-acervo"
    if "clima" in t:
        return "sec-clima"
    if "implica" in t or "lacuna" in t or "próximo" in t:
        return "sec-implica"
    return "sec-default"


def _html_tabela(rows: list[dict[str, Any]], max_cols: int = 8) -> str:
    if not rows:
        return ""
    keys: list[str] = []
    seen: set[str] = set()
    prefer = (
        "nm_urna", "nm_candidato", "qt_votos", "sg_partido", "vr_despesa", "vr_receita",
        "ds_sit_tot_turno", "ano", "sg_uf", "ds_cargo", "titulo", "fonte", "resumo",
    )
    for k in prefer:
        if any(k in row for row in rows) and k not in seen:
            seen.add(k)
            keys.append(k)
    for row in rows:
        for k in row:
            if k not in seen and not k.startswith("_"):
                seen.add(k)
                keys.append(k)
    keys = keys[:max_cols]
    parts = ["<div class='table-wrap'><table><thead><tr>"]
    parts.extend(f"<th>{_esc(k)}</th>" for k in keys)
    parts.append("</tr></thead><tbody>")
    for row in rows[:80]:
        parts.append("<tr>")
        parts.extend(f"<td>{_esc(str(row.get(k, '')))}</td>" for k in keys)
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    if len(rows) > 80:
        parts.append(f"<p class='table-note'>Exibindo 80 de {len(rows)} linhas. Exporte XLSX para o conjunto completo.</p>")
    return "".join(parts)


_REPORT_CSS = """
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Fraunces:opsz,wght@9..144,600;9..144,700&display=swap');
:root {
  --ink: #0c1222; --muted: #5c6478; --line: #e2e8f0;
  --brand: #0d4f4a; --brand-light: #e6f4f2;
  --acervo: #4338ca; --acervo-light: #eef2ff;
  --clima: #b45309; --clima-light: #fffbeb;
  --implica: #475569; --implica-light: #f1f5f9;
}
* { box-sizing: border-box; }
body {
  font-family: "DM Sans", system-ui, sans-serif;
  color: var(--ink); line-height: 1.65; margin: 0; background: #f4f6fa;
}
.page { max-width: 920px; margin: 0 auto; padding: 32px 20px 48px; }
.hero {
  background: linear-gradient(135deg, var(--brand) 0%, #1a7a72 100%);
  color: #fff; border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
  box-shadow: 0 16px 48px rgba(13,79,74,.22);
}
.hero-kicker { font-size: .72rem; font-weight: 700; letter-spacing: .12em; text-transform: uppercase; opacity: .85; }
.hero h1 { font-family: Fraunces, serif; font-size: 1.75rem; margin: 10px 0 8px; line-height: 1.2; }
.hero-meta { font-size: .88rem; opacity: .9; }
.trilhas { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 24px; }
.pill {
  display: inline-block; padding: 5px 12px; border-radius: 999px;
  font-size: .72rem; font-weight: 700; letter-spacing: .04em; text-transform: uppercase;
}
.pill-fato { background: var(--brand-light); color: var(--brand); }
.pill-acervo { background: var(--acervo-light); color: var(--acervo); }
.pill-clima { background: var(--clima-light); color: var(--clima); }
.sec {
  background: #fff; border: 1px solid var(--line); border-radius: 16px;
  padding: 24px 28px; margin-bottom: 18px; border-left: 4px solid var(--brand);
}
.sec-fato { border-left-color: var(--brand); }
.sec-acervo { border-left-color: var(--acervo); background: linear-gradient(90deg, var(--acervo-light) 0%, #fff 28%); }
.sec-clima { border-left-color: var(--clima); background: linear-gradient(90deg, var(--clima-light) 0%, #fff 28%); }
.sec-implica { border-left-color: var(--implica); background: linear-gradient(90deg, var(--implica-light) 0%, #fff 28%); }
.sec h2 {
  font-family: Fraunces, serif; font-size: 1.15rem; margin: 0 0 14px; color: var(--brand);
}
.sec-acervo h2 { color: var(--acervo); }
.sec-clima h2 { color: var(--clima); }
.sec-implica h2 { color: var(--implica); }
.sec h3 { font-size: 1rem; margin: 18px 0 8px; }
.sec h4 { font-size: .92rem; margin: 14px 0 6px; color: var(--muted); }
.sec p { margin: 0 0 12px; color: #334155; }
.sec ul { margin: 0 0 12px 1.2rem; color: #334155; }
.sec li { margin-bottom: 6px; }
.sec strong { color: var(--ink); }
.sec a { color: #1a5f8a; }
.appendix { margin-top: 32px; }
.appendix h2 {
  font-family: Fraunces, serif; font-size: 1.2rem; color: var(--brand);
  margin: 0 0 16px; padding-bottom: 10px; border-bottom: 2px solid var(--brand-light);
}
.appendix-block { background: #fff; border: 1px solid var(--line); border-radius: 16px; padding: 20px; margin-bottom: 16px; }
.appendix-block h3 { font-size: .95rem; margin: 0 0 12px; color: var(--muted); text-transform: uppercase; letter-spacing: .05em; }
.table-wrap { overflow-x: auto; }
table { border-collapse: collapse; width: 100%; font-size: .82rem; }
th, td { border: 1px solid var(--line); padding: 8px 10px; text-align: left; vertical-align: top; }
th { background: var(--brand-light); color: var(--brand); font-weight: 600; }
tr:nth-child(even) td { background: #f8fafc; }
.table-note { font-size: .78rem; color: var(--muted); margin-top: 8px; }
.footer {
  margin-top: 36px; padding-top: 20px; border-top: 1px solid var(--line);
  font-size: .82rem; color: var(--muted); text-align: center;
}
@media print {
  body { background: #fff; }
  .page { padding: 0; max-width: none; }
  .hero { box-shadow: none; }
  .sec { break-inside: avoid; }
}
"""


def exportar_html(
    dados: dict[str, Any] | None,
    conteudo_md: str,
    titulo: str = "Apura · Relatório",
) -> str:
    camadas = _linhas_por_camada(dados)
    gerado = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    secoes = _secoes_from_md(conteudo_md or "")

    parts = [
        "<!DOCTYPE html><html lang='pt-BR'><head><meta charset='utf-8'>",
        "<meta name='viewport' content='width=device-width, initial-scale=1'>",
        f"<title>{_esc(titulo)}</title>",
        f"<style>{_REPORT_CSS}</style></head><body><div class='page'>",
        "<header class='hero'>",
        "<div class='hero-kicker'>Inteligência Eleitoral Brasil · Apura</div>",
        f"<h1>{_esc(titulo)}</h1>",
        f"<p class='hero-meta'>Relatório analítico · Gerado em { _esc(gerado) }</p>",
        "</header>",
        "<div class='trilhas'>",
        "<span class='pill pill-fato'>Trilha A · Fato</span>",
        "<span class='pill pill-acervo'>Trilha B · Acervo</span>",
        "<span class='pill pill-clima'>Trilha C · Clima</span>",
        "</div>",
    ]

    for titulo_sec, corpo in secoes:
        cls = _classe_secao(titulo_sec)
        html_corpo = _md_block(corpo)
        if not html_corpo:
            continue
        parts.append(f"<section class='sec {cls}'>")
        parts.append(f"<h2>{_esc(titulo_sec)}</h2>")
        parts.append(html_corpo)
        parts.append("</section>")

    has_dados = any(camadas[k] for k in camadas)
    if has_dados:
        parts.append("<div class='appendix'>")
        parts.append("<h2>Anexo · dados consultados</h2>")
        for label, key in (
            ("Urna e contexto oficial", "fato"),
            ("Acervo (trechos indexados)", "acervo"),
            ("Clima (indícios)", "clima"),
        ):
            rows = camadas[key]
            if not rows:
                continue
            parts.append("<div class='appendix-block'>")
            parts.append(f"<h3>{_esc(label)} · {len(rows)} registro(s)</h3>")
            parts.append(_html_tabela(rows))
            parts.append("</div>")
        parts.append("</div>")

    parts.append(
        "<footer class='footer'>Inteligência Eleitoral Brasil · Fontes: TSE, IBGE, MDS, Câmara · "
        "Cifra só na Trilha A · Acervo e clima são contexto, não substituem urna</footer>"
    )
    parts.append("</div></body></html>")
    return "".join(parts)
